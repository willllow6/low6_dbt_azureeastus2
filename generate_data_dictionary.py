import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
MODEL_FILL    = PatternFill("solid", fgColor="2E75B6")   # mid blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
MODEL_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)

THIN  = Side(style="thin",  color="B8CCE4")
MED   = Side(style="medium", color="2E75B6")
CELL_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
SECTION_FONT  = Font(name="Calibri", bold=True, color="1F3864", size=11)
WRAP_FONT     = Font(name="Calibri", size=10)

def apply_header(cell, value):
    cell.value = value
    cell.font  = HEADER_FONT
    cell.fill  = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=MED, right=MED, top=MED, bottom=MED)

def apply_model_row(ws, row, model_name):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=model_name)
    cell.font  = MODEL_FONT
    cell.fill  = MODEL_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(left=MED, right=MED, top=MED, bottom=MED)

def apply_body_row(ws, row, values, alt=False):
    fill = ALT_ROW_FILL if alt else WHITE_FILL
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font   = BODY_FONT
        cell.fill   = fill
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 5))

# ── Entity glossary ────────────────────────────────────────────────────────────
# (entity_name, data_table, grain, description)
entities = [
    (
        "Tenant",
        "shared_penn__entries\nshared_penn__selections\nshared_penn__awards",
        "One row per tenant",
        "A sub-operator or regional partition within Penn. Penn is multi-tenant, meaning multiple operators can each run the squads game independently under the Penn platform. Every entry, selection, and award carries a tenant_id and tenant_name to identify which operator it belongs to.",
    ),
    (
        "Tournament",
        "shared_penn__entries\nshared_penn__selections",
        "One row per tournament\n(attribute only — no\ndedicated shared view)",
        "The top-level competition structure. A tournament groups a series of rounds and belongs to a single tenant. tournament_id and tournament_name are surfaced as attributes on entries and selections. Users do not enter a tournament directly — they enter individual rounds.",
    ),
    (
        "Contest (Round)",
        "shared_penn__entries\nshared_penn__selections\nshared_penn__awards",
        "One row per round",
        "The competitive unit a user enters. A round sits within a tournament and has a sequence number (e.g. Round 1, Round 2). Each round spans multiple reveal days — users build their squad by picking a player on each day. In the data, contest_id always refers to a round. round_sequence indicates the round's position within its tournament.",
    ),
    (
        "User",
        "shared_penn__users",
        "One row per user",
        "A player registered on the Penn platform. user_id is Penn's internal identifier. penn_user_id is the external SSO identifier used to link the user to Penn's operator platform (e.g. for prize delivery). A user may have a tier (loyalty level) and a platform (ios, android, web) recorded at registration.",
    ),
    (
        "Entry (User Round)",
        "shared_penn__entries",
        "One row per user\nper round",
        "A user's participation in a specific round. Created when a user joins a round. Tracks the user's status within that round, their tier at the time of entry, the prize outcome once the round resolves, and per-goal prize parameters. user_entry_number counts how many squads entries this user has ever made across all time (1 = their first ever squads entry).",
    ),
    (
        "Selection (Daily Pick / Slot)",
        "shared_penn__selections",
        "One row per user\nper reveal day\nper round",
        "A user's player pick for a single reveal day within their round entry. Each round has a fixed number of reveal days (e.g. 5 days = 5 slots per entry). On each reveal day, within the reveal window (reveal_start → reveal_end), a soccer player is revealed to the user. The user's slot then carries the revealed player's details, their rarity tier, and the exact reveal timestamp. A slot with a null player_id has not yet been revealed.",
    ),
    (
        "Award",
        "shared_penn__awards",
        "One row per award\n(one per qualifying\ngoal event)",
        "A monetary prize triggered by a qualifying in-game goal event. When a player in a user's squad scores a goal during a round, an award is created for that user's entry. Each award records the goal and player that triggered it, the prize amount and currency, and the delivery status. A single entry may earn multiple awards if multiple players in the squad score goals.",
    ),
]


def build_about_sheet(wb):
    ws = wb.create_sheet("About")
    ws.sheet_view.showGridLines = False

    # Column widths: Entity | Shared Views | Grain | Description
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 80

    # ── Title banner ──
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Penn Squads — Shared Data Dictionary")
    title_cell.font  = TITLE_FONT
    title_cell.fill  = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 30

    # ── Game overview ──
    ws.merge_cells("A2:D2")
    overview = (
        "Penn Squads is a squad-building fantasy game. Users enter rounds (contests) within a tournament and build a squad by picking one "
        "soccer player per reveal day over the course of the round. Awards (cash prizes) are paid per qualifying goal scored by a revealed player "
        "in the user's squad. Penn is multi-tenant: multiple operators each run independent instances of the game under their own tenant."
    )
    overview_cell = ws.cell(row=2, column=1, value=overview)
    overview_cell.font      = WRAP_FONT
    overview_cell.fill      = ALT_ROW_FILL
    overview_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    overview_cell.border    = Border(left=MED, right=MED, top=MED, bottom=MED)
    ws.row_dimensions[2].height = 52

    # ── Section heading ──
    ws.merge_cells("A3:D3")
    section_cell = ws.cell(row=3, column=1, value="Entity Glossary")
    section_cell.font      = SECTION_FONT
    section_cell.fill      = WHITE_FILL
    section_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 20

    # ── Entity table headers ──
    entity_headers = ["Entity", "Shared View(s)", "Grain", "Description"]
    entity_col_widths = [22, 34, 24, 80]
    for col, (hdr, _w) in enumerate(zip(entity_headers, entity_col_widths), start=1):
        apply_header(ws.cell(row=4, column=col), hdr)
    ws.row_dimensions[4].height = 22

    # ── Entity rows ──
    for row_idx, (entity, views, grain, desc) in enumerate(entities, start=5):
        alt = (row_idx % 2 == 1)
        fill = ALT_ROW_FILL if alt else WHITE_FILL
        row_h = max(16 * max(views.count("\n") + 1, grain.count("\n") + 1, 2), 16)

        for col, val in enumerate([entity, views, grain, desc], start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = fill
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=(1 if col == 1 else 0))
            if col == 1:
                cell.font = BOLD_FONT
            else:
                cell.font = WRAP_FONT
        ws.row_dimensions[row_idx].height = row_h

    # ── DAG flow diagram (text) ──
    dag_start_row = 5 + len(entities) + 1

    ws.merge_cells(start_row=dag_start_row, start_column=1, end_row=dag_start_row, end_column=4)
    dag_hdr = ws.cell(row=dag_start_row, column=1, value="Entity Relationships")
    dag_hdr.font      = SECTION_FONT
    dag_hdr.fill      = WHITE_FILL
    dag_hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[dag_start_row].height = 20

    dag_text = (
        "Tenant  ──▶  Tournament  ──▶  Contest (Round)  ──▶  Entry (User Round)  ──▶  Selection (Daily Pick)\n"
        "                                                             │\n"
        "                                                             └──▶  Award (per qualifying goal)"
    )
    ws.merge_cells(start_row=dag_start_row + 1, start_column=1, end_row=dag_start_row + 1, end_column=4)
    dag_cell = ws.cell(row=dag_start_row + 1, column=1, value=dag_text)
    dag_cell.font      = Font(name="Courier New", size=10)
    dag_cell.fill      = ALT_ROW_FILL
    dag_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    dag_cell.border    = Border(left=MED, right=MED, top=MED, bottom=MED)
    ws.row_dimensions[dag_start_row + 1].height = 50

    ws.freeze_panes = "A5"

# ── Data ───────────────────────────────────────────────────────────────────────
models = {
    "shared_penn__users": [
        ("user_id",              "VARCHAR",       "PK",  "No",  "Primary key. Unique identifier for the user in the Penn system."),
        ("penn_user_id",         "VARCHAR",       "",    "Yes", "Penn's external user identifier used to link to Penn's operator platform. Null for users not yet linked via SSO."),
        ("tier",                 "VARCHAR",       "",    "Yes", "User tier/loyalty level within the Penn platform."),
        ("platform",             "VARCHAR",       "",    "Yes", "Platform on which the user registered (e.g. ios, android, web)."),
        ("registration_date",    "DATE",          "",    "No",  "Date the user registered (UTC)."),
        ("registration_date_et", "DATE",          "",    "No",  "Date the user registered (Eastern Time)."),
        ("registered_at",        "TIMESTAMP_NTZ", "",   "No",  "Timestamp when the user registered (UTC)."),
        ("registered_at_et",     "TIMESTAMP_NTZ", "",   "No",  "Timestamp when the user registered (Eastern Time)."),
    ],
    "shared_penn__entries": [
        ("entry_id",             "VARCHAR",       "PK",  "No",  "Primary key. Unique identifier per user per round. One row per user per round."),
        ("user_id",              "VARCHAR",       "FK",  "No",  "Foreign key to the user record in the Penn system."),
        ("penn_user_id",         "VARCHAR",       "",    "Yes", "Penn's external user identifier. Null for users not yet linked via SSO."),
        ("contest_id",           "VARCHAR",       "FK",  "No",  "Foreign key to the round (contest). A round is the competitive unit a user enters."),
        ("tournament_id",        "VARCHAR",       "FK",  "No",  "Foreign key to the tournament. A tournament groups multiple rounds."),
        ("game_type",            "VARCHAR",       "",    "No",  "Game type. Always 'squads' for this view."),
        ("contest_name",         "VARCHAR",       "",    "Yes", "Display name of the round."),
        ("tournament_name",      "VARCHAR",       "",    "Yes", "Display name of the tournament."),
        ("contest_status",       "VARCHAR",       "",    "Yes", "Current status of the round (e.g. active, completed, cancelled)."),
        ("round_sequence",       "INTEGER",       "",    "Yes", "Sequence number of this round within its tournament (1 = first round)."),
        ("tenant_id",            "VARCHAR",       "FK",  "No",  "Foreign key to the tenant. Identifies the sub-operator partition."),
        ("tenant_name",          "VARCHAR",       "",    "Yes", "Display name of the tenant."),
        ("contest_starts_at",    "TIMESTAMP_NTZ", "",   "Yes", "Round start timestamp (UTC)."),
        ("contest_starts_at_et", "TIMESTAMP_NTZ", "",   "Yes", "Round start timestamp (Eastern Time)."),
        ("contest_start_date_et","DATE",          "",    "Yes", "Round start date (Eastern Time)."),
        ("user_tier",            "VARCHAR",       "",    "Yes", "User's tier at the time of entry."),
        ("entry_status",         "VARCHAR",       "",    "Yes", "Current status of the user's entry (e.g. active, completed)."),
        ("per_goal_amount",      "NUMBER(10,2)",  "",    "Yes", "Prize amount awarded per qualifying goal event. Null if not applicable."),
        ("per_goal_currency",    "VARCHAR",       "",    "Yes", "Currency for the per-goal prize amount (ISO 4217, e.g. USD)."),
        ("user_entry_number",    "INTEGER",       "",    "No",  "The user's nth entry in this game type across all time. 1 = first ever squads entry."),
        ("user_entry_type",      "VARCHAR",       "",    "No",  "Derived label: 'First Entry' if user_entry_number = 1, otherwise 'Returning Entry'."),
        ("entry_date",           "DATE",          "",    "No",  "Date of entry into the round (UTC)."),
        ("entered_at",           "TIMESTAMP_NTZ", "",   "No",  "Timestamp of entry into the round (UTC)."),
        ("entry_date_et",        "DATE",          "",    "No",  "Date of entry into the round (Eastern Time)."),
        ("entry_day_et",         "VARCHAR",       "",    "No",  "Day of the week of entry (Eastern Time, e.g. Monday)."),
        ("entry_hour_et",        "INTEGER",       "",    "No",  "Hour of entry (Eastern Time, 0–23)."),
        ("entered_at_et",        "TIMESTAMP_NTZ", "",   "No",  "Timestamp of entry into the round (Eastern Time)."),
        ("updated_at",           "TIMESTAMP_NTZ", "",   "No",  "Timestamp of the most recent update to the entry record."),
    ],
    "shared_penn__selections": [
        ("selection_id",    "VARCHAR",       "PK",  "No",  "Primary key. Unique identifier for the user's slot selection on a given reveal day."),
        ("entry_id",        "VARCHAR",       "FK",  "No",  "Foreign key to the user's entry (user_round). One entry spans multiple slots across reveal days."),
        ("user_id",         "VARCHAR",       "FK",  "No",  "Foreign key to the user record in the Penn system."),
        ("penn_user_id",    "VARCHAR",       "",    "Yes", "Penn's external user identifier. Null for users not yet linked via SSO."),
        ("contest_id",      "VARCHAR",       "FK",  "No",  "Foreign key to the round (contest)."),
        ("tournament_id",   "VARCHAR",       "FK",  "No",  "Foreign key to the tournament."),
        ("tenant_id",       "VARCHAR",       "FK",  "No",  "Foreign key to the tenant."),
        ("game_type",       "VARCHAR",       "",    "No",  "Game type. Always 'squads' for this view."),
        ("contest_name",    "VARCHAR",       "",    "Yes", "Display name of the round."),
        ("tournament_name", "VARCHAR",       "",    "Yes", "Display name of the tournament."),
        ("tenant_name",     "VARCHAR",       "",    "Yes", "Display name of the tenant."),
        ("day_index",       "INTEGER",       "",    "No",  "The reveal day index within the round (1 = first reveal day)."),
        ("status",          "VARCHAR",       "",    "Yes", "Current status of this slot (e.g. pending, revealed, locked)."),
        ("player_id",       "VARCHAR",       "FK",  "Yes", "Foreign key to the revealed soccer player. Null if the slot has not yet been revealed."),
        ("first_name",      "VARCHAR",       "",    "Yes", "Player's first name. Null if slot not yet revealed."),
        ("last_name",       "VARCHAR",       "",    "Yes", "Player's last name. Null if slot not yet revealed."),
        ("position",        "VARCHAR",       "",    "Yes", "Player's position (e.g. GK, DEF, MID, FWD). Null if slot not yet revealed."),
        ("number",          "INTEGER",       "",    "Yes", "Player's squad number. Null if slot not yet revealed."),
        ("birth_country",   "VARCHAR",       "",    "Yes", "Player's country of birth. Null if slot not yet revealed."),
        ("revealed_rarity", "VARCHAR",       "",    "Yes", "Rarity tier of the revealed player card (e.g. common, rare, epic). Null if not yet revealed."),
        ("revealed_at",     "TIMESTAMP_NTZ", "",   "Yes", "UTC timestamp when the player was revealed for this slot."),
        ("revealed_at_et",  "TIMESTAMP_NTZ", "",   "Yes", "Eastern Time timestamp when the player was revealed for this slot."),
        ("reveal_start",    "TIMESTAMP_NTZ", "",   "Yes", "UTC timestamp marking the start of the reveal window for this slot."),
        ("reveal_end",      "TIMESTAMP_NTZ", "",   "Yes", "UTC timestamp marking the end of the reveal window for this slot."),
        ("created_at",      "TIMESTAMP_NTZ", "",   "No",  "UTC timestamp when the slot record was created in the source system."),
    ],
    "shared_penn__awards": [
        ("award_id",           "VARCHAR",       "PK",  "No",  "Primary key. Unique identifier for the award record."),
        ("entry_id",           "VARCHAR",       "FK",  "No",  "Foreign key to the user's entry (user_round) this award is associated with."),
        ("user_id",            "VARCHAR",       "FK",  "No",  "Foreign key to the user record in the Penn system."),
        ("penn_user_id",       "VARCHAR",       "",    "Yes", "Penn's external user identifier used to process the payout. Null for users not yet linked via SSO."),
        ("user_tier",          "VARCHAR",       "",    "Yes", "User's tier at the time of the award."),
        ("contest_id",         "VARCHAR",       "FK",  "No",  "Foreign key to the round (contest) this award is associated with."),
        ("tenant_id",          "VARCHAR",       "FK",  "No",  "Foreign key to the tenant."),
        ("promotion_id",       "VARCHAR",       "",    "Yes", "Identifier of the promotion that triggered this award. Null if not promotion-linked."),
        ("goal_id",            "VARCHAR",       "",    "Yes", "Identifier of the goal event that triggered this award."),
        ("player_external_id", "VARCHAR",       "",    "Yes", "External identifier of the player associated with the goal event."),
        ("event_external_id",  "VARCHAR",       "",    "Yes", "External identifier of the match/event in which the goal occurred."),
        ("outcome_id",         "VARCHAR",       "",    "Yes", "Identifier of the prize outcome definition."),
        ("prize_amount",       "NUMBER(10,2)",  "",    "No",  "Monetary prize amount awarded."),
        ("currency",           "VARCHAR",       "",    "No",  "Currency of the prize amount (ISO 4217, e.g. USD)."),
        ("award_status",       "VARCHAR",       "",    "No",  "Current status of the award (e.g. pending, delivered, failed)."),
        ("platform",           "VARCHAR",       "",    "Yes", "Platform on which the award was processed (e.g. ios, android, web)."),
        ("delivered_at",       "TIMESTAMP_NTZ", "",   "Yes", "UTC timestamp when the award was successfully delivered to the user."),
        ("delivered_date_et",  "DATE",          "",    "Yes", "Eastern Time date when the award was delivered."),
        ("created_at",         "TIMESTAMP_NTZ", "",   "No",  "UTC timestamp when the award record was created in the source system."),
        ("updated_at",         "TIMESTAMP_NTZ", "",   "No",  "UTC timestamp of the most recent update to the award record."),
    ],
}

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── About sheet (first) ────────────────────────────────────────────────────────
build_about_sheet(wb)

# ── Index sheet ────────────────────────────────────────────────────────────────
idx = wb.create_sheet("Index")
idx.sheet_view.showGridLines = False
idx.column_dimensions["A"].width = 42
idx.column_dimensions["B"].width = 18

apply_header(idx.cell(row=1, column=1), "Shared View")
apply_header(idx.cell(row=1, column=2), "Row Count")
idx.row_dimensions[1].height = 22

for i, model_name in enumerate(models, start=2):
    idx.cell(row=i, column=1, value=model_name).font = BODY_FONT
    idx.cell(row=i, column=1).border = CELL_BORDER
    idx.cell(row=i, column=2, value=len(models[model_name])).font = BODY_FONT
    idx.cell(row=i, column=2).border = CELL_BORDER
    idx.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
    idx.cell(row=i, column=1).fill = fill
    idx.cell(row=i, column=2).fill = fill

# ── One sheet per model ────────────────────────────────────────────────────────
col_widths = [28, 16, 8, 10, 70]
col_headers = ["Column Name", "Data Type", "Key", "Nullable", "Description"]

for model_name, columns in models.items():
    ws = wb.create_sheet(model_name.replace("shared_penn__", ""))
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Model name banner
    apply_model_row(ws, 1, model_name)
    ws.row_dimensions[1].height = 20

    # Column headers
    for col, hdr in enumerate(col_headers, start=1):
        apply_header(ws.cell(row=2, column=col), hdr)
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, (col_name, dtype, key, nullable, desc) in enumerate(columns, start=3):
        alt = (row_idx % 2 == 1)
        apply_body_row(ws, row_idx, [col_name, dtype, key, nullable, desc], alt=alt)
        ws.row_dimensions[row_idx].height = 16

    ws.freeze_panes = "A3"

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = r"c:\Users\WillBreeden\low6_dbt_azureeastus2\penn_shared_data_dictionary.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
